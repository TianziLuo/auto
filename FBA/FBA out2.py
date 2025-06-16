from openpyxl import load_workbook
import xlwt
from datetime import datetime
import os
import time
import sys

#------- config --------
MAX_FILE_AGE = 30  # Save time in 30s
source_path = r"C:\Frank\3.1_FBA.xlsx"
template_path = r"C:\Template\出库.xlsx"

# Verify that the source file is fresh enough
try:
    file_age_seconds = time.time() - os.path.getmtime(source_path)
except FileNotFoundError:
    print(f"❌ Source file not found: {source_path}")
    sys.exit(1)

if file_age_seconds > MAX_FILE_AGE:
    print(
        f"❌ Source file is {int(file_age_seconds)} s old (> {MAX_FILE_AGE}s). Aborting."
    )
    sys.exit(1)

# Get today's date and generate filenames
today_str = datetime.today().strftime("%Y%m%d")
base_filename = f"出库_FBA_{today_str}_2"
downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
xlsx_path = os.path.join(downloads_path, f"{base_filename}.xlsx")
xls_path = os.path.join(downloads_path, f"{base_filename}.xls")

# Load source workbook and worksheet
source_wb = load_workbook(source_path, data_only=True)
source_ws = source_wb["FBA扣减库存"]

# Load template workbook and use the first sheet
template_wb = load_workbook(template_path)
template_ws = template_wb.active

# Define column and row ranges (E-H columns: 43-46, starting from row 3)
start_col = 43
end_col = 46
start_row = 3

# Find the last non-empty row in columns E-H
last_row = source_ws.max_row
while last_row >= start_row and all(source_ws.cell(row=last_row, column=col).value is None for col in range(start_col, end_col + 1)):
    last_row -= 1

# Start pasting data from row 2 in the template
paste_start_row = 2

# Copy data from source to template (map E-H → A-D)
for idx, row in enumerate(range(start_row, last_row + 1), start=paste_start_row):
    for col in range(start_col, end_col + 1):
        value = source_ws.cell(row=row, column=col).value
        template_ws.cell(row=idx, column=col - start_col + 1).value = value

# Save updated template as .xlsx file
template_wb.save(xlsx_path)
print(f"✅ Saved .xlsx to: {xlsx_path}")

# Reload the .xlsx file for xls conversion
wb_xlsx = load_workbook(xlsx_path, data_only=True)
ws_xlsx = wb_xlsx.active

# Create new .xls workbook
wb_xls = xlwt.Workbook()
ws_xls = wb_xls.add_sheet("Sheet1")

# Copy data cell by cell to .xls format
for row_idx, row in enumerate(ws_xlsx.iter_rows(values_only=True)):
    for col_idx, cell_value in enumerate(row):
        ws_xls.write(row_idx, col_idx, "" if cell_value is None else cell_value)

# Save as .xls file
wb_xls.save(xls_path)
print(f"✅ Also saved as .xls: {xls_path}")

# ---------- Clean up temporary .xlsx ----------
if os.path.exists(xlsx_path):
    try:
        os.remove(xlsx_path)
        print(f"🗑️ Deleted .xlsx file {xlsx_path}")
    except OSError as e:
        print(f"⚠️ Could not delete {xlsx_path}: {e}")
