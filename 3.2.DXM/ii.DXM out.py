from openpyxl import load_workbook
import xlwt
from datetime import datetime
import os

# Get today's date and generate filenames
base_filename = f"店小秘 更新库存"
downloads_path = r"C:\Frank\原始数据\店小秘+TP+订单+盘点"
xlsx_path = os.path.join(downloads_path, f"{base_filename}.xlsx")

# Source and template paths
source_path = r"C:\Frank\2.2_店小秘.xlsx"
template_path = r"C:\Template\店小秘 更新库存.xlsx"

# Load source workbook and worksheet
source_wb = load_workbook(source_path, data_only=True)
source_ws = source_wb["盘点"]

# Load template workbook and use the first sheet
template_wb = load_workbook(template_path)
template_ws = template_wb.active

# Define column and row ranges (E-H columns: 5-8, starting from row 3)
start_col = 1
end_col = 7
start_row = 2

# Find the last non-empty row in columns E-H
last_row = source_ws.max_row
while last_row >= start_row and all(source_ws.cell(row=last_row, column=col).value is None for col in range(start_col, end_col + 1)):
    last_row -= 1

# Start pasting data from row 2 in the template
paste_start_row = 2

# Copy data from source to template (map E-H → A-D)
for idx, row in enumerate(range(start_row, last_row + 1), start=paste_start_row):
    for col in range(start_col, end_col + 1):
        value = source_ws.cell(row=row, column=col).value
        template_ws.cell(row=idx, column=col - start_col + 1).value = value

# Save updated template as .xlsx file
template_wb.save(xlsx_path)
print(f"✅ Saved .xlsx to: {xlsx_path}")

