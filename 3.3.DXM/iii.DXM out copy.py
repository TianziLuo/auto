from openpyxl import load_workbook
from datetime import datetime
from pathlib import Path
import os

# 自动获取当前用户的下载目录
downloads_path = Path.home() / "Downloads"

# 设置保存文件名
base_filename = "店小秘 更新库存"
xlsx_path = downloads_path / f"{base_filename}.xlsx"

# 源文件和模板路径（仍为固定路径）
source_path = r"C:\Frank\2.2_店小秘.xlsx"
template_path = r"C:\Template\店小秘 更新库存.xlsx"

# 载入源工作簿与工作表
source_wb = load_workbook(source_path, data_only=True)
source_ws = source_wb["盘点"]

# 载入模板工作簿并使用第一个工作表
template_wb = load_workbook(template_path)
template_ws = template_wb.active

# 设置读取和写入的列行范围
start_col = 1  # 从A列（原E列）开始
end_col = 7    # 到G列（原H列）
start_row = 2  # 从第2行开始

# 找到E-H列的最后非空行
last_row = source_ws.max_row
while last_row >= start_row and all(source_ws.cell(row=last_row, column=col).value is None for col in range(start_col, end_col + 1)):
    last_row -= 1

# 开始粘贴到模板的第2行
paste_start_row = 2

# 从源复制数据到模板，E-H列→A-D列
for idx, row in enumerate(range(start_row, last_row + 1), start=paste_start_row):
    for col in range(start_col, end_col + 1):
        value = source_ws.cell(row=row, column=col).value
        template_ws.cell(row=idx, column=col - start_col + 1).value = value

# 保存到下载区
template_wb.save(xlsx_path)
print(f"✅ 文件已保存到下载区: {xlsx_path}")
